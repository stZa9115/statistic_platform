import openpyxl
from openpyxl.styles import Font, Alignment
import pandas as pd
def write_anova_postTest(posttest, file_path):
    pretest = pd.read_excel(file_path)
    groupNum = len(pretest)
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active


    sheet.cell(groupNum + 3, 1).value = "Post-hoc comparisons"
    sheet.cell(groupNum + 3, 1).font = Font(bold=True)
    sheet.cell(groupNum + 3, 1).alignment = Alignment(horizontal='center')
    sheet.merge_cells(start_row=groupNum + 3, start_column=1, end_row=groupNum + 3, end_column=5)


    for c in range(len(posttest.columns)):
        sheet.cell(groupNum + 4, c + 1).value = list(posttest.columns)[c]
        sheet.cell(groupNum + 4, c + 1).font = Font(bold=True)
        sheet.cell(groupNum + 4, c + 1).alignment = Alignment(horizontal='center')
    startPos = groupNum + 5
    for r in range(len(posttest)):
        colNum = 1
        for c in posttest.columns:
            sheet.cell(startPos + r, colNum).value = posttest.loc[r, c]
            sheet.cell(startPos + r, colNum).alignment = Alignment(horizontal='center')
            colNum += 1
    wb.save(file_path)
    wb.save('./check/full.xlsx')
if __name__ == "__main__":
    # Example usage
    post_df = pd.read_excel('./check/posttest.xlsx')
    write_anova_postTest(post_df, './check/pretest.xlsx')